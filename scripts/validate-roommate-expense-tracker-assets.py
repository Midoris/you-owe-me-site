#!/usr/bin/env python3
"""Validate the roommate expense tracker workbook, CSV, image, and page links."""

from __future__ import annotations

import csv
import importlib.util
import json
import os
import re
import sys
from pathlib import Path
from zipfile import ZipFile


if importlib.util.find_spec("PIL") is None:
    bundled_python = (
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "python"
        / "bin"
        / "python3"
    )
    if bundled_python.is_file() and Path(sys.executable).resolve() != bundled_python.resolve():
        os.execv(str(bundled_python), [str(bundled_python), *sys.argv])

from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "downloads" / "roommate-expense-tracker-template.xlsx"
CSV_PATH = ROOT / "downloads" / "roommate-expense-tracker-template.csv"
IMAGE_PATH = (
    ROOT
    / "images"
    / "tools"
    / "roommate-expense-tracker-template"
    / "roommate-expense-tracker-spreadsheet-hero.webp"
)
PAGE_PATH = ROOT / "tools" / "roommate-expense-tracker-template" / "index.html"
SHEETS = ["Start Here", "Setup", "Expenses", "Repayments", "Summary", "Example", "Settings"]
CSV_HEADERS = [
    "Date",
    "Entry type",
    "Expense or repayment",
    "Category",
    "Amount",
    "Paid by or From",
    "To",
    "Split method",
    "Included roommates",
    "Agreed shares",
    "Notes",
]


def settlement_transfers(names: list[str], positions: list[float], valid: bool = True):
    if not valid or abs(sum(positions)) > 0.01:
        return None
    balances = [round(value, 2) for value in positions]
    transfers = []
    for _ in range(max(0, len(names) - 1)):
        debtor = next((index for index, value in enumerate(balances) if value < -0.01), None)
        creditor = next((index for index, value in enumerate(balances) if value > 0.01), None)
        if debtor is None or creditor is None:
            break
        amount = round(min(-balances[debtor], balances[creditor]), 2)
        transfers.append((names[debtor], names[creditor], amount))
        balances[debtor] = round(balances[debtor] + amount, 2)
        balances[creditor] = round(balances[creditor] - amount, 2)
    assert all(abs(value) <= 0.01 for value in balances), balances
    return transfers


def validate_workbook() -> None:
    formulas = load_workbook(WORKBOOK_PATH, data_only=False)
    values = load_workbook(WORKBOOK_PATH, data_only=True)
    assert formulas.sheetnames == SHEETS, formulas.sheetnames

    with ZipFile(WORKBOOK_PATH) as archive:
        names = archive.namelist()
        assert "xl/vbaProject.bin" not in names
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        assert "externalReference" not in workbook_xml

    start = formulas["Start Here"]
    assert start["A1"].value == "Roommate Expense Tracker Spreadsheet"
    assert "positive position" in start["A12"].value
    assert "does not decide fair rent" in start["A20"].value

    setup = formulas["Setup"]
    assert setup["A1"].value == "Household setup"
    assert [setup.cell(row, 2).value for row in range(9, 15)] == [None] * 6
    assert [setup.cell(row, 3).value for row in range(9, 15)] == [0] * 6
    setup_formula = setup["A17"].value
    for state in [
        "Ready: opening positions balance to 0.00.",
        "Opening positions do not balance. Check that the total is 0.00.",
        "Add at least two roommate names.",
        "Roommate names must be unique.",
    ]:
        assert state in setup_formula

    expenses = formulas["Expenses"]
    assert [expenses.cell(4, column).value for column in range(1, 7)] == [
        "Date",
        "Expense",
        "Category",
        "Amount",
        "Paid by",
        "Split method",
    ]
    assert expenses["S4"].value == "Notes"
    assert expenses["Z4"].value == "Share check"
    assert all(expenses.cell(4, column).data_type == "f" for column in range(7, 13))
    assert all(expenses.cell(4, column).data_type == "f" for column in range(13, 19))
    assert all(expenses.cell(4, column).data_type == "f" for column in range(20, 26))
    assert all(expenses.cell(row, 1).value is None for row in range(5, 205))
    assert all(expenses.cell(row, 19).value is None for row in range(5, 205))
    assert all(expenses.cell(row, 20).data_type == "f" for row in (5, 6, 50, 204))
    assert all(expenses.cell(row, 26).data_type == "f" for row in (5, 6, 50, 204))
    assert expenses.freeze_panes == "A5"
    expense_validations = "\n".join(
        f"{validation.sqref} {validation.formula1}"
        for validation in expenses.data_validations.dataValidation
    )
    for expected in ["C5:C204", "D5:D204", "E5:E204", "F5:F204", "G5:L204", "M5:R204"]:
        assert expected in expense_validations
    check_formula = expenses["Z5"].value
    for state in [
        "Ready",
        "Choose at least one included roommate.",
        "Choose who paid.",
        "Enter an amount greater than 0.",
        "Custom shares are short by ",
        "Custom shares are over by ",
        "Enter custom shares that add up to the expense amount.",
    ]:
        assert state in check_formula

    repayments = formulas["Repayments"]
    assert [repayments.cell(4, column).value for column in range(1, 7)] == [
        "Date",
        "From",
        "To",
        "Amount",
        "Note",
        "Check",
    ]
    assert all(repayments.cell(row, 1).value is None for row in range(5, 205))
    assert all(repayments.cell(row, 6).data_type == "f" for row in (5, 6, 50, 204))
    repayment_formula = repayments["F5"].value
    for state in [
        "Ready",
        "Choose two different roommates.",
        "Choose who sent and who received the repayment.",
        "Enter an amount greater than 0.",
    ]:
        assert state in repayment_formula

    summary = formulas["Summary"]
    assert [summary.cell(4, column).value for column in range(1, 9)] == [
        "Roommate",
        "Opening position",
        "Paid toward expenses",
        "Assigned expense share",
        "Repayments sent",
        "Repayments received",
        "Current position",
        "Status",
    ]
    assert summary["G5"].value == "=IF(A5=\"\",\"\",B5+C5-D5+E5-F5)"
    assert all(values["Summary"].cell(row, 1).value in (None, "") for row in range(5, 11))
    assert all(summary.column_dimensions[column].hidden for column in "JKLMNOPQRS")
    record_formula = summary["B17"].value
    for state in [
        "Ready: current positions balance to 0.00.",
        "Resolve the highlighted setup, expense, or repayment checks before settling.",
        "Current positions do not balance to 0.00. Check the opening positions and expense shares.",
    ]:
        assert state in record_formula
    settlement_state = summary["A29"].value
    assert "No settlement transfer is needed." in settlement_state
    assert "Settlement suggestions are unavailable until the current positions balance to 0.00." in settlement_state
    assert all(summary.cell(row, column).data_type == "f" for row in range(23, 28) for column in range(1, 4))
    assert values["Summary"]["B17"].value == "Resolve the highlighted setup, expense, or repayment checks before settling."

    example = values["Example"]
    assert example["B20"].value == 1980
    assert example["B21"].value == 660
    exact = {
        "Maya": (1800, 660, 0, 200, 940),
        "Alex": (60, 660, 200, 0, -400),
        "Sam": (120, 660, 0, 0, -540),
    }
    for row in range(16, 19):
        name = example.cell(row, 1).value
        assert tuple(example.cell(row, column).value for column in range(2, 7)) == exact[name]
    assert (example["A25"].value, example["B25"].value, example["C25"].value) == ("Alex", "Maya", 400)
    assert (example["A26"].value, example["B26"].value, example["C26"].value) == ("Sam", "Maya", 540)
    assert example["F24"].value == 0

    settings = formulas["Settings"]
    assert [settings.cell(4, column).value for column in range(1, 5)] == [
        "Categories",
        "Split methods",
        "Yes or no",
        "Currencies",
    ]
    assert [settings.cell(row, 4).value for row in range(5, 16)] == [
        "USD",
        "EUR",
        "GBP",
        "CAD",
        "AUD",
        "NZD",
        "JPY",
        "CHF",
        "SGD",
        "THB",
        "Other",
    ]

    # Pure-model checks mirror the workbook's documented settlement engine.
    assert settlement_transfers(["A", "B", "C"], [900, -400, -500]) == [
        ("B", "A", 400),
        ("C", "A", 500),
    ]
    assert settlement_transfers(["A", "B", "C", "D"], [300, 200, -250, -250]) == [
        ("C", "A", 250),
        ("D", "A", 50),
        ("D", "B", 200),
    ]
    assert settlement_transfers(["A", "B", "C"], [0, 0, 0]) == []
    assert settlement_transfers(["Maya", "Alex", "Sam"], [90, -90, 0]) == [
        ("Alex", "Maya", 90),
    ]
    assert settlement_transfers(["A", "B"], [10, -9]) is None
    assert settlement_transfers(["A", "B"], [10, -10], valid=False) is None


def validate_csv() -> None:
    with CSV_PATH.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == CSV_HEADERS
    assert len(rows) == 4
    assert all(len(row) == len(CSV_HEADERS) for row in rows)
    assert {row[5] for row in rows[1:]} == {"Maya", "Alex"}
    assert "does not calculate balances" in rows[1][-1]
    assert any("," in row[8] for row in rows[1:])


def validate_image() -> None:
    with Image.open(IMAGE_PATH) as image:
        assert image.size == (1536, 1024)
        assert image.format == "WEBP"


def validate_page() -> None:
    html = PAGE_PATH.read_text(encoding="utf-8")
    assert html.count("<h1") == 1
    assert "Roommate Expense Tracker Spreadsheet | Free Template" in html
    assert '<link rel="canonical" href="https://you-owe-me.com/tools/roommate-expense-tracker-template/" />' in html
    assert html.count("<!-- best-next-step:start -->") == 1
    assert html.count("<!-- best-next-step:end -->") == 1
    assert "roommate-expense-tracker-spreadsheet-hero.webp" in html
    assert "Roommate expense tracker spreadsheet with shared bills, repayments, and a monthly settle-up summary." in html
    for filename in ["roommate-expense-tracker-template.xlsx", "roommate-expense-tracker-template.csv"]:
        links = re.findall(rf'<a\b[^>]*href="[^"]*{re.escape(filename)}"[^>]*>', html, re.I | re.S)
        assert links, filename
        assert all("download" in link for link in links), filename
    assert "https://apps.apple.com/us/app/loan-tracker-you-owe-me/id1147058670?ppid=18039f2b-da9e-4d5f-9ba1-b60f117ecf12" in html
    assert 'data-track-location="roommate_expense_spreadsheet_product_app_store"' in html
    scripts = re.findall(r'<script type="application/ld\+json">\s*(.*?)\s*</script>', html, re.S)
    documents = [json.loads(script) for script in scripts]
    nodes = [node for document in documents for node in document.get("@graph", [document])]
    faq = next(node for node in nodes if node.get("@type") == "FAQPage")
    assert len(faq["mainEntity"]) == 7
    assert not any(node.get("@type") in {"SoftwareApplication", "WebApplication", "HowTo"} for node in nodes)


def main() -> None:
    validate_workbook()
    validate_csv()
    validate_image()
    validate_page()
    print("validated roommate workbook, CSV, image, formulas, settlement cases, and page links")


if __name__ == "__main__":
    main()
