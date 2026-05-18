from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = ROOT / "downloads"

HEADERS = [
    "Date",
    "Family member",
    "Entry type",
    "Category",
    "What it was for",
    "Paid by",
    "Expense amount",
    "Repayment amount",
    "Balance change",
    "Running balance",
    "Status",
    "Review / due date",
    "Receipt / reference",
    "Notes",
]

ENTRY_TYPES = [
    "Expense paid for family",
    "Repayment received",
    "Adjustment",
]

CATEGORIES = [
    "Utilities",
    "Subscription",
    "Groceries",
    "Pharmacy / care purchase",
    "Online order",
    "Household item",
    "Travel",
    "Meal",
    "Caregiving support",
    "School / children",
    "Repayment",
    "Other",
]

STATUS_OPTIONS = [
    "Needs reimbursement",
    "Partly repaid",
    "Repaid",
    "Review later",
    "Not reimbursable",
]

REVIEW_STATUS_OPTIONS = [
    "Done",
    "Needs follow-up",
    "Review later",
]

PAID_BY_OPTIONS = [
    "Me",
    "Family member",
    "Sibling",
    "Partner",
    "Other",
]

EXAMPLE_ROWS = [
    [date(2026, 5, 1), "Mom", "Expense paid for family", "Utilities", "Monthly internet bill", "Me", 48, None, "Needs reimbursement", date(2026, 5, 15), "Internet May", "Recurs monthly"],
    [date(2026, 5, 3), "Dad", "Expense paid for family", "Pharmacy / care purchase", "Pharmacy pickup", "Me", 23, None, "Needs reimbursement", date(2026, 5, 15), "Receipt 1042", "Keep details general"],
    [date(2026, 5, 5), "Mom", "Expense paid for family", "Online order", "Household supplies", "Me", 37, None, "Needs reimbursement", date(2026, 5, 15), "Order May 5", "Ordered online"],
    [date(2026, 5, 8), "Mom", "Repayment received", "Repayment", "Partial repayment", "Mom", None, 40, "Partly repaid", None, "Bank transfer", "Partial repayment"],
    [date(2026, 5, 10), "Family shared", "Expense paid for family", "Subscription", "Shared cloud storage", "Me", 12, None, "Needs reimbursement", date(2026, 6, 1), "Monthly", "Repeats monthly"],
    [date(2026, 5, 11), "Sister", "Expense paid for family", "Groceries", "Groceries for family dinner", "Me", 64, None, "Review later", date(2026, 5, 20), "Grocery receipt", "Split later"],
    [date(2026, 5, 12), "Sister", "Repayment received", "Repayment", "Repaid grocery share", "Sister", None, 32, "Partly repaid", None, "Cash", "Half repaid"],
    [date(2026, 5, 14), "Dad", "Repayment received", "Repayment", "Full repayment", "Dad", None, 23, "Repaid", None, "Bank transfer", "Dad fully repaid"],
    [date(2026, 5, 15), "Mom", "Expense paid for family", "Subscription", "Streaming subscription", "Me", 15, None, "Needs reimbursement", date(2026, 6, 1), "Subscription", "Monthly subscription"],
    [date(2026, 5, 16), "Mom", "Adjustment", "Other", "Removed duplicate entry", "Me", 0, 15, "Review later", None, "Correction", "Corrected duplicate subscription"],
    [date(2026, 5, 18), "Family shared", "Repayment received", "Repayment", "Shared subscription repayment", "Family member", None, 12, "Repaid", None, "Bank transfer", "Shared subscription settled"],
]


def add_defined_name(workbook, name, target):
    workbook.defined_names.add(DefinedName(name, attr_text=target))


def style_header(row):
    header_fill = PatternFill("solid", fgColor="AFE67E")
    border = Border(bottom=Side(style="thin", color="D8E2D1"))
    for cell in row:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="18212B")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def set_standard_columns(sheet):
    widths = {
        "A": 13,
        "B": 18,
        "C": 24,
        "D": 23,
        "E": 28,
        "F": 16,
        "G": 16,
        "H": 18,
        "I": 16,
        "J": 16,
        "K": 19,
        "L": 17,
        "M": 22,
        "N": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def apply_tracker_format(sheet, max_row):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max_row}"
    sheet.row_dimensions[1].height = 30
    style_header(sheet[1])
    set_standard_columns(sheet)

    for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=14):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E4E9EF"))

    for row in range(2, max_row + 1):
        sheet[f"A{row}"].number_format = "yyyy-mm-dd"
        sheet[f"L{row}"].number_format = "yyyy-mm-dd"
        for col in ("G", "H", "I", "J"):
            sheet[f"{col}{row}"].number_format = '$#,##0.00;-$#,##0.00;""'

    warm_fill = PatternFill("solid", fgColor="FFF4C2")
    settled_fill = PatternFill("solid", fgColor="EAF7DF")
    negative_fill = PatternFill("solid", fgColor="EEF2F6")
    sheet.conditional_formatting.add(f"J2:J{max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=warm_fill))
    sheet.conditional_formatting.add(f"J2:J{max_row}", CellIsRule(operator="equal", formula=["0"], fill=settled_fill))
    sheet.conditional_formatting.add(f"J2:J{max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill))


def add_validations(workbook, sheet, max_row):
    validations = [
        ("C", "=EntryTypes"),
        ("D", "=Categories"),
        ("F", "=PaidByOptions"),
        ("K", "=StatusOptions"),
    ]
    for col, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{col}2:{col}{max_row}")


def build_workbook():
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    start = wb.active
    start.title = "Start Here"
    tracker = wb.create_sheet("Tracker")
    example = wb.create_sheet("Example Tracker")
    summary = wb.create_sheet("Summary")
    monthly = wb.create_sheet("Monthly Review")
    settings = wb.create_sheet("Settings")

    start.sheet_view.showGridLines = False
    start.merge_cells("A1:H1")
    start["A1"] = "Family Reimbursement Tracker Template"
    start["A1"].fill = PatternFill("solid", fgColor="AFE67E")
    start["A1"].font = Font(size=20, bold=True, color="18212B")
    start["A1"].alignment = Alignment(vertical="center")
    start.row_dimensions[1].height = 34
    start["A3"] = "Use this workbook to track money you pay for parents, siblings, relatives, shared family costs, bills, subscriptions, purchases, repayments, and remaining balances."
    start.merge_cells("A3:H3")
    start["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions = [
        "Use the Tracker sheet for your own entries.",
        "Add expenses when you pay for something.",
        "Add repayments as separate rows.",
        "Check the Summary sheet to see current balances.",
        "Use the Example Tracker sheet if you want to see how partial repayments work.",
        "Use the Monthly Review sheet to prepare one calm weekly or monthly summary.",
        "Edit categories in the Settings sheet if needed.",
    ]
    start["A5"] = "How to use it"
    start["A5"].font = Font(bold=True, size=14, color="18212B")
    for index, text in enumerate(instructions, start=1):
        row = 5 + index
        start[f"A{row}"] = index
        start[f"B{row}"] = text
        start[f"A{row}"].font = Font(bold=True, color="355428")
        start[f"B{row}"].alignment = Alignment(wrap_text=True)
    notes = [
        ("Privacy note", "Family money can be sensitive. Keep this file somewhere private and avoid storing unnecessary personal or medical details."),
        ("Limitation note", "This template is for personal organization only. It is not legal, tax, accounting, or medical advice."),
        ("You Owe Me note", "If this spreadsheet becomes hard to maintain, You Owe Me can help you keep running balances, repayment history, recurring entries, reminders, and PDF statements in one place."),
    ]
    row = 14
    for title, text in notes:
        start[f"A{row}"] = title
        start[f"A{row}"].font = Font(bold=True, color="18212B")
        start[f"B{row}"] = text
        start[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 2
    start.column_dimensions["A"].width = 18
    for column in range(2, 9):
        start.column_dimensions[get_column_letter(column)].width = 20

    settings.sheet_view.showGridLines = False
    settings.append(["Entry types", "Categories", "Status options", "Paid by options", "Review status options"])
    style_header(settings[1])
    max_settings_rows = max(len(ENTRY_TYPES), len(CATEGORIES), len(STATUS_OPTIONS), len(PAID_BY_OPTIONS), len(REVIEW_STATUS_OPTIONS))
    for index in range(max_settings_rows):
        settings.append([
            ENTRY_TYPES[index] if index < len(ENTRY_TYPES) else None,
            CATEGORIES[index] if index < len(CATEGORIES) else None,
            STATUS_OPTIONS[index] if index < len(STATUS_OPTIONS) else None,
            PAID_BY_OPTIONS[index] if index < len(PAID_BY_OPTIONS) else None,
            REVIEW_STATUS_OPTIONS[index] if index < len(REVIEW_STATUS_OPTIONS) else None,
        ])
    settings["A16"] = "You can edit these lists if your family uses different categories."
    settings["A16"].font = Font(italic=True, color="5F6D7A")
    settings.merge_cells("A16:E16")
    for column in ("A", "B", "C", "D", "E"):
        settings.column_dimensions[column].width = 26

    add_defined_name(wb, "EntryTypes", "Settings!$A$2:$A$4")
    add_defined_name(wb, "Categories", "Settings!$B$2:$B$13")
    add_defined_name(wb, "StatusOptions", "Settings!$C$2:$C$6")
    add_defined_name(wb, "PaidByOptions", "Settings!$D$2:$D$6")
    add_defined_name(wb, "ReviewStatusOptions", "Settings!$E$2:$E$4")

    tracker.append(HEADERS)
    for row in range(2, 302):
        tracker[f"I{row}"] = f'=IF(C{row}="Expense paid for family",G{row},IF(C{row}="Repayment received",-H{row},IF(C{row}="Adjustment",G{row}-H{row},"")))'
        tracker[f"J{row}"] = f'=IF(B{row}="","",SUMIFS($I$2:I{row},$B$2:B{row},B{row}))'
    tracker["G1"].comment = Comment("Expense amount is used for expenses.", "You Owe Me")
    tracker["H1"].comment = Comment("Repayment amount is used for repayments.", "You Owe Me")
    tracker["I1"].comment = Comment("Balance change and Running balance are calculated automatically.", "You Owe Me")
    tracker["J1"].comment = Comment("Running balance is calculated separately for each family member.", "You Owe Me")
    apply_tracker_format(tracker, 301)
    add_validations(wb, tracker, 301)

    example.append(HEADERS)
    for idx, row_values in enumerate(EXAMPLE_ROWS, start=2):
        example[f"A{idx}"] = row_values[0]
        example[f"B{idx}"] = row_values[1]
        example[f"C{idx}"] = row_values[2]
        example[f"D{idx}"] = row_values[3]
        example[f"E{idx}"] = row_values[4]
        example[f"F{idx}"] = row_values[5]
        example[f"G{idx}"] = row_values[6]
        example[f"H{idx}"] = row_values[7]
        example[f"I{idx}"] = f'=IF(C{idx}="Expense paid for family",G{idx},IF(C{idx}="Repayment received",-H{idx},IF(C{idx}="Adjustment",G{idx}-H{idx},"")))'
        example[f"J{idx}"] = f'=IF(B{idx}="","",SUMIFS($I$2:I{idx},$B$2:B{idx},B{idx}))'
        example[f"K{idx}"] = row_values[8]
        example[f"L{idx}"] = row_values[9]
        example[f"M{idx}"] = row_values[10]
        example[f"N{idx}"] = row_values[11]
    apply_tracker_format(example, 12)

    summary.sheet_view.showGridLines = False
    summary.append(["Family member", "Total expenses", "Total repayments", "Current balance", "Last review note"])
    style_header(summary[1])
    family_names = ["Mom", "Dad", "Sister", "Brother", "Family shared", "Other"]
    for index in range(2, 32):
        summary[f"A{index}"] = family_names[index - 2] if index - 2 < len(family_names) else None
        summary[f"B{index}"] = f'=IF($A{index}="","",SUMIFS(Tracker!$G:$G,Tracker!$B:$B,$A{index}))'
        summary[f"C{index}"] = f'=IF($A{index}="","",SUMIFS(Tracker!$H:$H,Tracker!$B:$B,$A{index}))'
        summary[f"D{index}"] = f'=IF($A{index}="","",B{index}-C{index})'
        for col in ("B", "C", "D"):
            summary[f"{col}{index}"].number_format = '$#,##0.00;-$#,##0.00;""'
    total_row = 33
    summary[f"A{total_row}"] = "Total open balance"
    summary[f"A{total_row}"].font = Font(bold=True)
    summary[f"D{total_row}"] = "=SUM(D2:D31)"
    summary[f"D{total_row}"].font = Font(bold=True)
    summary[f"D{total_row}"].number_format = '$#,##0.00;-$#,##0.00;""'
    summary["A35"] = "Edit the family member names in column A to match your situation."
    summary["A35"].font = Font(italic=True, color="5F6D7A")
    summary.merge_cells("A35:E35")
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = "A1:E31"
    for column, width in {"A": 22, "B": 18, "C": 18, "D": 18, "E": 34}.items():
        summary.column_dimensions[column].width = width
    summary.conditional_formatting.add("D2:D31", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFF4C2")))

    monthly.sheet_view.showGridLines = False
    monthly.merge_cells("A1:F1")
    monthly["A1"] = "Monthly Family Reimbursement Review"
    monthly["A1"].fill = PatternFill("solid", fgColor="AFE67E")
    monthly["A1"].font = Font(size=18, bold=True, color="18212B")
    monthly["A1"].alignment = Alignment(vertical="center")
    monthly.row_dimensions[1].height = 32
    monthly["A3"] = "Review period"
    monthly["C3"] = "Review date"
    monthly["E3"] = "Prepared by"
    for cell in ("A3", "C3", "E3"):
        monthly[cell].font = Font(bold=True, color="18212B")
    for cell in ("B3", "D3", "F3"):
        monthly[cell].fill = PatternFill("solid", fgColor="FFFFFF")
        monthly[cell].border = Border(bottom=Side(style="thin", color="9AA7B3"))
    monthly.merge_cells("A5:F5")
    monthly["A5"] = "Current balances to review"
    monthly["A5"].font = Font(bold=True, size=13, color="18212B")
    monthly["A5"].fill = PatternFill("solid", fgColor="E4E4B8")
    monthly["A6"] = "Family member"
    monthly["B6"] = "Current balance"
    monthly["C6"] = "Review status"
    monthly["D6"] = "Follow-up date"
    monthly["E6"] = "Summary message"
    monthly["F6"] = "Notes"
    style_header(monthly[6])
    for row in range(7, 37):
        source_row = row - 5
        monthly[f"A{row}"] = f'=IF(Summary!A{source_row}="","",Summary!A{source_row})'
        monthly[f"B{row}"] = f'=IF(A{row}="","",Summary!D{source_row})'
        monthly[f"B{row}"].number_format = '$#,##0.00;-$#,##0.00;""'
        monthly[f"D{row}"].number_format = "yyyy-mm-dd"
        for col in range(1, 7):
            monthly.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            monthly.cell(row, col).border = Border(bottom=Side(style="hair", color="E4E9EF"))
    status_validation = DataValidation(type="list", formula1="=ReviewStatusOptions", allow_blank=True)
    monthly.add_data_validation(status_validation)
    status_validation.add("C7:C36")
    monthly["A39"] = "Monthly review checklist"
    monthly["A39"].font = Font(bold=True, size=13, color="18212B")
    monthly["A39"].fill = PatternFill("solid", fgColor="E4E4B8")
    monthly.merge_cells("A39:F39")
    monthly["A40"] = "Checklist item"
    monthly["B40"] = "Status"
    monthly["C40"] = "Notes"
    monthly.merge_cells("C40:F40")
    style_header(monthly[40])
    checklist_items = [
        "Recurring bills added",
        "Partial repayments recorded",
        "Current balances checked",
        "Old items reviewed",
        "One clear summary prepared if needed",
        "Next review date chosen",
    ]
    for index, item in enumerate(checklist_items, start=41):
        monthly[f"A{index}"] = item
        monthly[f"A{index}"].alignment = Alignment(wrap_text=True)
        monthly.merge_cells(start_row=index, start_column=3, end_row=index, end_column=6)
        for col in range(1, 7):
            monthly.cell(index, col).border = Border(bottom=Side(style="hair", color="E4E9EF"))
    checklist_validation = DataValidation(type="list", formula1="=ReviewStatusOptions", allow_blank=True)
    monthly.add_data_validation(checklist_validation)
    checklist_validation.add("B41:B46")
    monthly["A49"] = "One calm summary"
    monthly["A49"].font = Font(bold=True, size=13, color="18212B")
    monthly["A50"] = "Use this space to draft one short family reimbursement summary from the current balance instead of rebuilding the story from memory."
    monthly["A50"].alignment = Alignment(wrap_text=True)
    monthly.merge_cells("A50:F50")
    monthly["A52"] = "Draft summary:"
    monthly["A52"].font = Font(bold=True, color="18212B")
    monthly.merge_cells("B52:F55")
    monthly["B52"].alignment = Alignment(wrap_text=True, vertical="top")
    monthly["B52"].fill = PatternFill("solid", fgColor="FFFFFF")
    for row in range(52, 56):
        for col in range(2, 7):
            monthly.cell(row, col).border = Border(bottom=Side(style="hair", color="C8D2DC"))
    monthly.freeze_panes = "A6"
    monthly.auto_filter.ref = "A6:F36"
    for column, width in {"A": 24, "B": 16, "C": 18, "D": 16, "E": 24, "F": 34}.items():
        monthly.column_dimensions[column].width = width
    monthly.conditional_formatting.add("B7:B36", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFF4C2")))

    for sheet in wb.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.35
        sheet.page_margins.right = 0.35
        sheet.page_margins.top = 0.55
        sheet.page_margins.bottom = 0.55

    return wb


def build_csv():
    csv_path = DOWNLOADS / "family-reimbursement-tracker-template.csv"
    csv_path.write_text(",".join(HEADERS) + "\n", encoding="utf-8")


def paragraph(text, style):
    return Paragraph(text, style)


def draw_pdf_page_background(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(colors.white)
    canvas.rect(0, 0, doc.pagesize[0], doc.pagesize[1], stroke=0, fill=1)
    canvas.restoreState()


def build_pdf():
    pdf_path = DOWNLOADS / "family-reimbursement-tracker-printable.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(letter),
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.42 * inch,
        bottomMargin=0.42 * inch,
        title="Family Reimbursement Tracker",
    )

    styles = getSampleStyleSheet()
    title = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=24, leading=28, alignment=TA_LEFT, textColor=colors.HexColor("#18212B"))
    subtitle = ParagraphStyle("Subtitle", parent=styles["BodyText"], fontName="Helvetica", fontSize=12.5, leading=17, textColor=colors.HexColor("#344054"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontName="Helvetica", fontSize=10.5, leading=15, textColor=colors.HexColor("#344054"))
    small = ParagraphStyle("Small", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=12.5, textColor=colors.HexColor("#5F6D7A"))
    table_header = ParagraphStyle("TableHeader", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=7.8, leading=9.1, alignment=TA_CENTER, textColor=colors.HexColor("#18212B"))

    elements = []
    elements.append(paragraph("Family Reimbursement Tracker", title))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(paragraph("Use this printable log to record money paid for family members, repayments received, and what is still open.", subtitle))
    elements.append(Spacer(1, 0.28 * inch))
    instruction_rows = [
        ["1", "Write each expense as soon as you pay it."],
        ["2", "Record repayments as separate rows."],
        ["3", "Update the remaining balance after each repayment."],
        ["4", "Review the log weekly or monthly."],
        ["5", "Print extra copies of the blank log page if one month needs more rows."],
    ]
    instruction_table = Table(instruction_rows, colWidths=[0.35 * inch, 6.8 * inch])
    instruction_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#AFE67E")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#18212B")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("LEADING", (0, 0), (-1, -1), 15),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8E2D1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E9EF")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(instruction_table)
    elements.append(Spacer(1, 0.28 * inch))
    quick_reference = Table(
        [
            ["Quick reference", ""],
            ["Expense", "Money you paid for a family member or shared family cost."],
            ["Repaid", "Money sent back to you. Record repayments as separate rows."],
            ["Remaining balance", "What is still open after each expense or repayment."],
        ],
        colWidths=[1.65 * inch, 5.85 * inch],
    )
    quick_reference.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#E4E4B8")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#18212B")),
        ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.4),
        ("LEADING", (0, 0), (-1, -1), 12.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8E2D1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E9EF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(quick_reference)
    elements.append(Spacer(1, 0.22 * inch))
    elements.append(paragraph("<b>Privacy note:</b> Keep private family or care details brief. Use a receipt/reference number instead of writing sensitive details.", body))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(paragraph("This printable log is for personal organization only. It is not legal, tax, accounting, or medical advice.", small))
    elements.append(PageBreak())

    log_meta = Table(
        [["Month / period", "", "Family member or group", "", "Reviewed on", ""]],
        colWidths=[1.05 * inch, 1.45 * inch, 1.55 * inch, 1.45 * inch, 1.05 * inch, 1.45 * inch],
    )
    log_meta.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#344054")),
        ("LINEBELOW", (1, 0), (1, 0), 0.7, colors.HexColor("#9AA7B3")),
        ("LINEBELOW", (3, 0), (3, 0), 0.7, colors.HexColor("#9AA7B3")),
        ("LINEBELOW", (5, 0), (5, 0), 0.7, colors.HexColor("#9AA7B3")),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(paragraph("Blank printable log", title))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(log_meta)
    elements.append(Spacer(1, 0.1 * inch))

    headers = [
        paragraph("Date", table_header),
        paragraph("Family member", table_header),
        paragraph("What it was for", table_header),
        paragraph("Expense", table_header),
        paragraph("Repaid", table_header),
        paragraph("Remaining<br/>balance", table_header),
        paragraph("Notes", table_header),
    ]
    blank_rows = [["", "", "", "", "", "", ""] for _ in range(15)]
    log_table = Table(
        [headers] + blank_rows,
        colWidths=[0.78 * inch, 1.15 * inch, 2.35 * inch, 0.78 * inch, 0.78 * inch, 1.45 * inch, 2.05 * inch],
        repeatRows=1,
    )
    log_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#AFE67E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#18212B")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C8D2DC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFCFE")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(log_table)
    elements.append(PageBreak())

    elements.append(paragraph("Monthly review checklist", title))
    elements.append(Spacer(1, 0.12 * inch))
    checklist = [
        ["[  ]", "Recurring bills added"],
        ["[  ]", "Partial repayments recorded"],
        ["[  ]", "Remaining balance checked"],
        ["[  ]", "Old items reviewed"],
        ["[  ]", "One clear summary prepared if needed"],
    ]
    checklist_table = Table(checklist, colWidths=[0.55 * inch, 5.5 * inch])
    checklist_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E4E4B8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("LEADING", (0, 0), (-1, -1), 16),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8E2D1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E9EF")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(checklist_table)
    elements.append(Spacer(1, 0.25 * inch))
    elements.append(paragraph("If the printable log becomes hard to maintain, move the records into the Excel template or You Owe Me.", body))
    elements.append(Spacer(1, 0.18 * inch))
    elements.append(paragraph("<b>One calm summary:</b> Use this space to draft one short message from the current balance instead of rebuilding the story from memory.", body))
    elements.append(Spacer(1, 0.08 * inch))
    summary_box = Table([[""]], colWidths=[7.3 * inch], rowHeights=[1.1 * inch])
    summary_box.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D2DC")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FBFCFE")),
    ]))
    elements.append(summary_box)

    doc.build(elements, onFirstPage=draw_pdf_page_background, onLaterPages=draw_pdf_page_background)


def main():
    DOWNLOADS.mkdir(parents=True, exist_ok=True)
    build_csv()
    build_pdf()
    workbook = build_workbook()
    workbook.save(DOWNLOADS / "family-reimbursement-tracker-template.xlsx")


if __name__ == "__main__":
    main()
