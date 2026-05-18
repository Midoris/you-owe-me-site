import csv
import json
import re
import urllib.parse
import xml.etree.ElementTree as ET
from html import unescape
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]


def main():
    html = (ROOT / "tools/family-reimbursement-tracker-template/index.html").read_text(encoding="utf-8")
    sitemap = (ROOT / "sitemap.xml").read_text(encoding="utf-8")
    robots = (ROOT / "robots.txt").read_text(encoding="utf-8")
    llms = (ROOT / "llms.txt").read_text(encoding="utf-8")
    workbook_path = ROOT / "downloads/family-reimbursement-tracker-template.xlsx"
    csv_path = ROOT / "downloads/family-reimbursement-tracker-template.csv"
    pdf_path = ROOT / "downloads/family-reimbursement-tracker-printable.pdf"

    wb = load_workbook(workbook_path, data_only=False)
    assert wb.sheetnames == ["Start Here", "Tracker", "Example Tracker", "Summary", "Monthly Review", "Settings"], wb.sheetnames

    tracker = wb["Tracker"]
    assert [tracker.cell(1, col).value for col in range(1, 15)] == [
        "Date",
        "Family member",
        "Entry type",
        "Category",
        "What it was for",
        "Paid by",
        "Expense amount",
        "Repayment amount",
        "Balance change",
        "Running balance",
        "Status",
        "Review / due date",
        "Receipt / reference",
        "Notes",
    ]
    assert tracker["I2"].value == '=IF(C2="Expense paid for family",G2,IF(C2="Repayment received",-H2,IF(C2="Adjustment",G2-H2,"")))'
    assert tracker["J2"].value == '=IF(B2="","",SUMIFS($I$2:I2,$B$2:B2,B2))'
    for row in (2, 3, 50, 301):
        assert tracker[f"I{row}"].value == f'=IF(C{row}="Expense paid for family",G{row},IF(C{row}="Repayment received",-H{row},IF(C{row}="Adjustment",G{row}-H{row},"")))'
        assert tracker[f"J{row}"].value == f'=IF(B{row}="","",SUMIFS($I$2:I{row},$B$2:B{row},B{row}))'
    assert tracker.freeze_panes == "A2"
    assert tracker.auto_filter.ref == "A1:N301"
    validations = "\n".join(str(validation.formula1) for validation in tracker.data_validations.dataValidation)
    assert "=EntryTypes" in validations
    assert "=Categories" in validations
    assert "=PaidByOptions" in validations
    assert "=StatusOptions" in validations
    for column in ("G", "H", "I", "J"):
        assert "$" in tracker[f"{column}2"].number_format

    example = wb["Example Tracker"]
    assert example.max_row >= 12
    example_rows = [
        tuple(example.cell(row, col).value for col in range(1, 15))
        for row in range(2, 13)
    ]
    assert any(row[1] == "Mom" and row[3] == "Utilities" for row in example_rows)
    assert any(row[2] == "Repayment received" and row[4] == "Partial repayment" for row in example_rows)
    assert any(row[2] == "Repayment received" and row[4] == "Full repayment" for row in example_rows)
    assert any(row[3] == "Subscription" for row in example_rows)
    expected_balances = {
        2: 48,
        3: 23,
        4: 85,
        5: 45,
        6: 12,
        7: 64,
        8: 32,
        9: 0,
        10: 60,
        11: 45,
        12: 0,
    }
    balances = {}
    for row in range(2, 13):
        assert example[f"I{row}"].value
        assert example[f"J{row}"].value
        member = example[f"B{row}"].value
        entry_type = example[f"C{row}"].value
        expense = example[f"G{row}"].value or 0
        repayment = example[f"H{row}"].value or 0
        if entry_type == "Expense paid for family":
            change = expense
        elif entry_type == "Repayment received":
            change = -repayment
        else:
            change = expense - repayment
        balances[member] = balances.get(member, 0) + change
        assert balances[member] == expected_balances[row], (row, balances[member], expected_balances[row])

    summary = wb["Summary"]
    assert summary["B2"].value == '=IF($A2="","",SUMIFS(Tracker!$G:$G,Tracker!$B:$B,$A2))'
    assert summary["C2"].value == '=IF($A2="","",SUMIFS(Tracker!$H:$H,Tracker!$B:$B,$A2))'
    assert summary["D2"].value == '=IF($A2="","",B2-C2)'
    assert summary["D33"].value == "=SUM(D2:D31)"
    assert summary["A33"].value == "Total open balance"

    monthly = wb["Monthly Review"]
    assert monthly["A7"].value == '=IF(Summary!A2="","",Summary!A2)'
    assert monthly["B7"].value == '=IF(A7="","",Summary!D2)'
    assert monthly["A41"].value == "Recurring bills added"
    assert monthly["A45"].value == "One clear summary prepared if needed"

    with csv_path.open(newline="", encoding="utf-8") as handle:
        headers = next(csv.reader(handle))
    assert headers == [
        "Date",
        "Family member",
        "Entry type",
        "Category",
        "What it was for",
        "Paid by",
        "Expense amount",
        "Repayment amount",
        "Balance change",
        "Running balance",
        "Status",
        "Review / due date",
        "Receipt / reference",
        "Notes",
    ]

    pdf = PdfReader(str(pdf_path))
    assert len(pdf.pages) == 3
    assert all((page.extract_text() or "").strip() for page in pdf.pages)
    page_two_text = " ".join((pdf.pages[1].extract_text() or "").split())
    assert "Remaining balance" in page_two_text

    with ZipFile(workbook_path) as archive:
        names = archive.namelist()
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        assert "xl/vbaProject.bin" not in names

    for download in re.findall(r'href="../../downloads/([^"]+)', html):
        assert (ROOT / "downloads" / download).exists(), download

    assert html.count("<h1") == 1
    assert '<link rel="canonical" href="https://you-owe-me.com/tools/family-reimbursement-tracker-template/" />' in html
    assert "<title>Family Reimbursement Tracker Template | Track Money Paid for Family</title>" in html
    assert '<meta property="og:title" content="Family Reimbursement Tracker Template" />' in html
    assert '<meta name="twitter:card" content="summary_large_image" />' in html
    assert "Download a free family reimbursement tracker template for money paid for parents, siblings, relatives, bills, subscriptions, purchases, repayments, and remaining balances." in html
    assert "Download a free Excel, CSV, and printable PDF tracker for family reimbursements, parent bills, subscriptions, repayments, and remaining balances." in html
    assert "Download Excel workbook (.xlsx)" in html
    assert "Download CSV template (.csv)" in html
    assert "Download printable PDF log (.pdf)" in html
    assert 'aria-label="Download the Family Reimbursement Tracker Excel workbook"' in html
    assert 'aria-label="Download the Family Reimbursement Tracker CSV template"' in html
    assert 'aria-label="Download the Family Reimbursement Tracker printable PDF log"' in html
    assert '<figcaption class="family-mini-sheet" aria-hidden="true">' in html
    assert '<table class="family-example-table">' in html
    assert "<caption>Example family reimbursement log with expenses, repayments, balance changes, and running balances.</caption>" in html
    assert "<thead>" in html
    assert "<tbody>" in html
    assert '<th scope="col">Date</th>' in html
    for outbound in [
        "/solutions/family-reimbursement-tracker/",
        "/solutions/app-to-track-money-owed/",
        "/tools/polite-payback-reminder-generator/",
        "/tools/repayment-reminder-text-examples/",
        "/tools/split-expense-calculator/",
        "/tools/",
        "/blog/how-to-keep-track-of-money-between-family-members/",
        "/features/",
    ]:
        assert f'href="{outbound}"' in html, outbound
    for inbound_page in [
        "tools/index.html",
        "solutions/family-reimbursement-tracker/index.html",
        "blog/how-to-keep-track-of-money-between-family-members/index.html",
        "blog/index.html",
        "solutions/index.html",
        "solutions/app-to-track-money-owed/index.html",
        "tools/repayment-reminder-text-examples/index.html",
        "tools/polite-payback-reminder-generator/index.html",
        "features/index.html",
    ]:
        assert "/tools/family-reimbursement-tracker-template/" in (ROOT / inbound_page).read_text(encoding="utf-8")

    scripts = re.findall(r'<script type="application/ld\+json">\s*(.*?)\s*</script>', html, re.S)
    faq_schema = None
    for script in scripts:
        document = json.loads(script)
        for node in document.get("@graph", [document]):
            if isinstance(node, dict) and node.get("@type") == "FAQPage":
                faq_schema = node
    assert faq_schema
    visible_faq = []
    for match in re.finditer(r"<details[^>]*>\s*<summary>(.*?)</summary>\s*<p>(.*?)</p>", html, re.S):
        question = unescape(re.sub(r"<.*?>", "", match.group(1)).strip())
        answer = unescape(re.sub(r"<.*?>", "", match.group(2)).strip())
        visible_faq.append((question, " ".join(answer.split())))
    schema_faq = [
        (item["name"], " ".join(item["acceptedAnswer"]["text"].split()))
        for item in faq_schema["mainEntity"]
    ]
    assert visible_faq[-len(schema_faq):] == schema_faq

    url = "https://you-owe-me.com/tools/family-reimbursement-tracker-template/"
    assert sitemap.count(url) == 1
    assert "family-reimbursement-tracker-template.xlsx" not in sitemap
    assert "family-reimbursement-tracker-template.csv" not in sitemap
    assert "family-reimbursement-tracker-printable.pdf" not in sitemap
    ET.fromstring(sitemap)
    assert "Sitemap: https://you-owe-me.com/sitemap.xml" in robots
    assert not re.search(r"Disallow:\s*/(tools|downloads)/?", robots, re.I)
    assert "Family Reimbursement Tracker Template" in llms

    broken_links = []
    for page in ROOT.glob("**/index.html"):
        content = page.read_text(encoding="utf-8", errors="ignore")
        for raw_href in re.findall(r'<a\b[^>]*\shref=["\']([^"\']+)["\']', content, re.I):
            if raw_href.startswith(("#", "mailto:", "tel:", "javascript:")):
                continue
            parsed = urllib.parse.urlsplit(unescape(raw_href))
            if parsed.scheme in ("http", "https") and parsed.netloc not in ("you-owe-me.com", "www.you-owe-me.com"):
                continue
            path = parsed.path if parsed.scheme else raw_href.split("#", 1)[0].split("?", 1)[0]
            if not path:
                continue
            target = ROOT / path.lstrip("/") if path.startswith("/") else page.parent / path
            if path.endswith("/") or not target.suffix:
                target = target / "index.html"
            if not target.exists():
                broken_links.append((str(page.relative_to(ROOT)), raw_href))
    assert not broken_links, broken_links[:8]

    assert "Open in Google Sheets" not in html
    assert "Family reimbursement tracker template shown as a calm spreadsheet workspace." in html
    assert "Monthly family reimbursement review with receipts and balance cards." in html

    print("validated workbook/csv/pdf/html downloads")


if __name__ == "__main__":
    main()
